import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dateutil import parser

# =====================================================
# FIXED FIELD SCHEMA
# =====================================================
FIELDS = [
    "Complaint ID",
    "Date Filed",
    "Date Accepted",
    "Time Accepted",
    "Complainant Name",
    "Email",
    "Mobile Number",
    "State",
    "District",
    "Cybercrime Type",
    "Sub-Category",
    "Platform",
    "Source Bank",
    "Amount Lost",
    "Transaction Count",
    "Date Range",
    "Complaint Status",
    "FIR Status",
    "Investigation Status"
]

# =====================================================
# LABEL MAP
# =====================================================
LABELS = {
    "Complaint ID": ["acknowledgement number"],
    "Date Filed": ["complaint date"],
    "Cybercrime Type": ["category of complaint"],
    "Sub-Category": ["sub category of complaint"],
    "Complainant Name": ["name"],
    "Email": ["email", "userid"],
    "Mobile Number": ["mobile"],
    "District": ["district"],
    "State": ["state"],
    "Amount Lost": ["total fraudulent amount"]
}

# =====================================================
# PDF TEXT EXTRACTION
# =====================================================
def extract_text_from_pdf(path):
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += "\n" + t
    return text

# =====================================================
# NORMALIZATION
# =====================================================
def clean_value(val):
    return re.sub(r"\s+", " ", val).strip(" :-")

def normalize_date(val):
    try:
        return parser.parse(val, dayfirst=True).strftime("%d/%m/%Y")
    except:
        return ""

def normalize_amount(val):
    num = re.sub(r"[^\d]", "", val)
    return f"₹{int(num):,}" if num.isdigit() else ""

# =====================================================
# MAIN FIELD EXTRACTION
# =====================================================
def extract_main_fields(text):
    data = {f: "" for f in FIELDS}
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    for line in lines:
        low = line.lower()
        for field, keys in LABELS.items():
            for key in keys:
                if low.startswith(key):
                    value = line[len(key):]
                    data[field] = clean_value(value)

    # Email fallback
    if not data["Email"]:
        m = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", text)
        if m:
            data["Email"] = m.group(0)

    # Accepted Date + Time
    m = re.search(
        r"complaint accepted date\s*(\d{1,2}/\d{1,2}/\d{4})\s*(\d{1,2}:\d{2}:\d{2}\s*[AP]M)",
        text,
        re.I
    )
    if m:
        data["Date Accepted"] = normalize_date(m.group(1))
        data["Time Accepted"] = m.group(2)

    # Normalize
    data["Date Filed"] = normalize_date(data["Date Filed"])
    data["Amount Lost"] = normalize_amount(data["Amount Lost"])

    # =================================================
    # TRANSACTION COUNT + DATE RANGE
    # =================================================
    dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", text)
    txn_dates = []
    for d in dates:
        try:
            txn_dates.append(parser.parse(d, dayfirst=True))
        except:
            pass

    data["Transaction Count"] = str(len(txn_dates)) if txn_dates else ""

    if txn_dates:
        data["Date Range"] = (
            min(txn_dates).strftime("%d/%m/%Y")
            + " - " +
            max(txn_dates).strftime("%d/%m/%Y")
        )

    # =================================================
    # STATUS
    # =================================================
    T = text.upper()
    if "COMPLAINT ACCEPTED" in T:
      data["Complaint Status"] = "ACCEPTED"
    elif "COMPLAINT REJECTED" in T:
      data["Complaint Status"] = "REJECTED"
    elif "COMPLAINT" not in T:
     data["Complaint Status"] = "NOT FILLED"
    else:
        data["Complaint Status"] = "PENDING"
    if "FIR" in T:
      data["FIR Status"] = "FILED"
    elif "FIR" not in T:
        data["FIR Status"] = "NOT FILLED"
    else:
       data["FIR Status"] = "NOT FILED"
    data["Investigation Status"] = (
        "CLOSED" if "CLOSED" in T else
        "ONGOING" if "UNDER PROCESS" in T else
        "NOT STARTED"
    )

    # =================================================
    # FINAL NULL FILL (CRITICAL)
    # =================================================
    for k in data:
        if not data[k]:
            data[k] = "NULL"

    return data

# =====================================================
# PROCESS PDF
# =====================================================
def process_pdf(pdf_path):
    text = extract_text_from_pdf(pdf_path)
    return {"MAIN_FIELDS": [extract_main_fields(text)]}

# =====================================================
# EXCEL FORMATTING
# =====================================================
def format_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    ws.freeze_panes = "A2"

# =====================================================
# SAVE EXCEL
# =====================================================
def save_consolidated_excel(all_data, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "MAIN_FIELDS"

    ws.append(["PDF"] + FIELDS)

    for i, data in enumerate(all_data, 1):
        row = data["MAIN_FIELDS"][0]
        ws.append([f"PDF_{i}"] + [row[f] for f in FIELDS])

    format_worksheet(ws)
    wb.save(out_path)
